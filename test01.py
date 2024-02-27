import os, sys
import openpyxl

from numpy import dot 
from numpy.linalg import norm 

from openai import OpenAI
from dotenv import load_dotenv

from unstructured.partition.auto import partition

SENTENCE="のりたまとは何ですか？"
# Wikipediaの「のりたま」の記事のURL
DATA_URL="https://ja.wikipedia.org/wiki/%E3%81%AE%E3%82%8A%E3%81%9F%E3%81%BE"


def create_vector(client:OpenAI, embedding_model_name, sentence):
    # ベクトルの作成
    res = client.embeddings.create(
        model=embedding_model_name,
        input=[sentence]
    )
    return res.data[0].embedding

def test_similarity(embedding_model_name, sentence, text_list, print_result=False):

    client = OpenAI(
        api_key=os.getenv('OPENAI_API_KEY')
    )
   
    # ベクトルの作成
    v1 = create_vector(client, embedding_model_name, sentence)

    result_list = []

    for i in range(len(text_list)):
        # iを001, 002, 003のように0埋め3桁で表示
        num = str(i+1).zfill(3)
        text = text_list[i]
        v = create_vector(client, embedding_model_name, text)
        # コサイン類似度の計算
        cos_sim = dot(v1, v) / (norm(v1) * norm(v)) 
        if print_result:
            print(f'case {num}: {cos_sim}: \n{text}')
        result_list.append({
            "case": num,
            "cos_sim": cos_sim,
            "text": text
        })
    return result_list

def load_text(url, chunk_size=1000):
    elements = partition(url=url)
    text_list = []
    text = ""
    for el in elements:
        text += str(el)
        # print(text)
        if len(text) > chunk_size:
            text_list.append(text)
            text = ""
    if len(text) > 0:
        text_list.append(text)

    return text_list

def save_result(output_file_path, result_list):
    # Excelファイルの書き込み
    book = openpyxl.Workbook()
    sheet = book.worksheets[0]
    for i in range(len(result_list)):
        row = i+1
        case = result_list[i]["case"]
        cos_sim = result_list[i]["cos_sim"]
        text = result_list[i]["text"]

        sheet.cell(row=row, column=1, value=case)
        sheet.cell(row=row, column=2, value=cos_sim)
        sheet.cell(row=row, column=3, value=text)
    
    book.save(output_file_path)    

def do_test(sentence, text_list, embedding_model_name):

    # Wikipediaの「のりたま」の記事を500文字ごとに分割してリストに格納
    # 各テキストと「のりたまとは何ですか？」との類似度を計算

    # .envファイルの読み込み
    load_dotenv()

    result_list = test_similarity(embedding_model_name, sentence, text_list, print_result=False)

    output_excel_name = "result_" + os.path.basename(sys.argv[0]).replace(".py", ".xlsx")
    save_result(output_excel_name, result_list)

    for item in result_list:
        print(f"case {item['case']}: {item['cos_sim']}")

def create_test_data(url, chunk_size=500):
    text_list = load_text(url, chunk_size)

    return text_list

if __name__ == "__main__":

    test_data_list = create_test_data(DATA_URL)
    do_test(SENTENCE, test_data_list, "text-embedding-3-small")

